#!/usr/bin/env python

import os
import sys
import re
import openpyxl as excel
import csv

from glob import glob
from argparse import ArgumentParser

Worksheet = excel.worksheet.worksheet.Worksheet
# from lib.chart import Chart

DESC = {'cons': 'Early conception',
        'child_helpline': 'Child Helpline number',
        'freedom': "Freedoms",
        "prof": "Professions",
        "like_games": "Liking games",
        "aspirations": "Sharing aspirations at home",
        "legal_marriage_age": "Legal marriage age",
        "serving_guests": "Serving guests",
        "social": "Social expectations",
        "financial_decisions": "Financial decisions at home",
        "games": "Playing games",
        "chores": "Household chores"}


def add_chart(root: Worksheet, type_: str) -> None:
    chart = Chart(root)

    def get_metadata() -> tuple[str, str, str, str, str]:
        data_ref: str | None = None
        x_axis_ref: str | None = None
        title: str | None = None
        x_axis_title: str | None = None
        y_axis_title: str | None = None

        if re.search(r'^(games|prof|social)', type_):
            data_ref = chart.data_ref(min_row=1, min_col=3)
            x_axis_ref = chart.x_axis_ref(min_row=2, max_col=2)
            title = 'Professions'
            x_axis_title = 'Profession with responses'
            y_axis_title = '% of respondents'

        return title, x_axis_title, y_axis_title, data_ref, x_axis_ref 

        # elif re.search(r'^()')

    title, x_axis_title, y_axis_title, data_ref, x_axis_ref = get_metadata()
    chart.bar_chart(title, x_axis_title, y_axis_title, data_ref, x_axis_ref)

def read_csv(file: str) -> list[list[str]]:
    with open(file) as fh:
        return [row for row in csv.reader(fh)]

def match_files(files: list[str], pat: str='.*') -> list[str]:
    res = []
    pat = re.compile(pat, re.I)

    for f in files:
        if pat.search(f): res.append(f)

    print(files, res)
    return res


def find_files(district: str, pat: str='.*') -> tuple[str, dict[str, list[str]]]:
    res = {}
    src = f'output/{district}'
    sheetdirs = glob(f'{src}/*')
    sheetdirs = [x for x in sheetdirs if os.path.isdir(x)]
    sheetdirs = sorted(sheetdirs)

    for col in sheetdirs:
        files = match_files(glob(f'{col}/*.csv'), pat)
        res[os.path.basename(col)] = sorted(files)

    return (district, res)


def create_excel(files: tuple[str, dict[str, list[str]]],
                 dest: str | None=None) -> str:
    district, files = files[0], files[1]
    wb = excel.Workbook() 
    dest = f'output/{district}.xlsx' if not dest else dest

    for sheet, csv_files in files.items():
        root: excel.worksheet.worksheet.Worksheet = wb.create_sheet(DESC[sheet])
        linenum: int = 1
        csv_files_len = len(csv_files)

        for i, src in enumerate(csv_files):
            data = read_csv(src)
            title = os.path.basename(src).replace('.csv', '')

            match title:
                case 'master':
                    title = f'Total respondents ({DESC[sheet]})'
                case 'master-pct':
                    title = f'% of respondents ({DESC[sheet]})'
                case 'master-Boy':
                    title = f'Boys\' responses ({DESC[sheet]})'
                case 'master-Girl':
                    title = f'Girls\' responses ({DESC[sheet]})'
                case 'master-Boy-pct':
                    title = f'% of Boy responses ({DESC[sheet]})'
                case 'master-Girl-pct':
                    title = f'% of Girl responses ({DESC[sheet]})'
                
            root.append([title])
            title_cell = root.cell(row=linenum, column=1)
            title_cell.style = 'Title'
            [root.append(row) for row in data]

            if i != csv_files_len - 1:
                root.append([''])

            linenum += len(data) + 2

    del wb['Sheet']
    wb.save(dest)
    return dest

create_excel(find_files(sys.argv[1], 'master'))
